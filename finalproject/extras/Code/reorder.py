#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Outputs a version where the phrases are re-ordered in SVO order
TODO Change the order within the phrase and improve reordering
"""
import openpyxl
import argparse
import codecs
import sys
from collections import defaultdict
from itertools import izip, islice
PARSER = argparse.ArgumentParser(description="Reorder a set of sentences")
PARSER.add_argument("-t", type=str, default="data/train", help="training data prefix")
PARSER.add_argument("-d", type=str, default="data/dev", help="dev data prefix")
PARSER.add_argument("-w", type=str, default="es", help="spanish file")
PARSER.add_argument("-p", type=str, default="data/postags.xlsx", help="tag file")
args = PARSER.parse_args()
sys.stdout = codecs.getwriter('utf-8')(sys.stdout)
sys.stdin = codecs.getreader('utf-8')(sys.stdin)
def sentences():
        with open(combine(args.d,args.w)) as f:
            for pair in f:
                yield [sentence.strip().split() for sentence in pair.split(' ||| ')]
def combine(a, b): return '%s.%s' % (a, b)
if __name__ == '__main__':
    if args.t:
        def utf8read(file): return codecs.open(file, 'r', 'utf-8')
        wb = openpyxl.load_workbook(args.p)
        parsed = wb.get_sheet_by_name('dev')
        sentence_no = 1
        order = []
        sentences = defaultdict(defaultdict)
        properties = []
        rows = int(parsed.get_highest_row())
        columns = int(parsed.get_highest_column())
        word_props = defaultdict(list)
        for i in range(1, rows+1):
            if parsed.cell(column = 1, row = i).value == 1:
                if i != 1:
                    sentences[sentence_no] = word_props
                    sentence_no +=1
                    word_props = defaultdict(list)
            word_props[i] =  [parsed.cell(column = 2, row = i).value, parsed.cell(column = 4, row = i).value, parsed.cell(column = 5, row = i).value,parsed.cell(column = 8, row = i).value]
        subjects = ["SUBJ"]
        final_sent = ""
        verb_phrase = ""
        subj_phrase = ""
        obj_phrase = ""
        noun_phrase = ""
        remains = ""

        flags = [False,False,False]
        objects = ["DO","IO","OBLC"]
        verbs = ["v"]
        for i in range(1, sentence_no):
            final_sent = ""
            verb_phrase = ""
            subj_phrase = ""
            obj_phrase = ""
            noun_phrase = ""
            remains = ""
            flags = [False,False,False]
            temp_phrase = []
            words = sentences[i]
            for pos in words:
                if words[pos][0] != None:
                    if (type(words[pos][0]) == long) | (type(words[pos][0]) == float) | (type(words[pos][0]) == int):
                        word = unicode(str(words[pos][0]),"utf-8")
                    else:
                        word = words[pos][0]
                    temp_phrase.append(word)
                if (words[pos][3] in subjects) & (flags[0] == False):
                    subj_phrase = ' '.join(temp_phrase)
                    flags[0] = True
                    temp_phrase = []
                elif (words[pos][3] in objects)&(flags[1]==False):
                    obj_phrase = ' '.join(temp_phrase)
                    flags[1] = True
                    temp_phrase = []
                elif (words[pos][1] in verbs)&(flags[2]==False):
                    verb_phrase = ' '.join(temp_phrase)
                    flags[2] = True
                    temp_phrase = []
                else:
                    remains = ' '.join(temp_phrase)
            final_sent = subj_phrase+' '+verb_phrase+' '+obj_phrase+' '+remains
            print final_sent