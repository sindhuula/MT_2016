#!/usr/bin/env python
# -*- coding: utf-8 -*-
import codecs
from collections import defaultdict
import openpyxl
import sys

wb = openpyxl.load_workbook("sample.xlsx")
parsed = wb.get_sheet_by_name('first_1000')
sys.stdout = codecs.getwriter('utf-8')(sys.stdout)
sys.stdin = codecs.getreader('utf-8')(sys.stdin)
if __name__ == '__main__':
    word_no = 0
    sentence_no = 0
    order = []
    sentences = defaultdict(str)
    flags = [False, False, False]
    subjects = ["SUBJ"]
    objects = ["DO","IO","OBLC"]
    verbs = ["AUX"]
    SOV = []
    SVO = []
    VSO = []
    VOS = []
    OVS = []
    OSV = []
    other = []
    string = []
    rows = int( parsed.get_highest_row())
    columns = int(parsed.get_highest_column())
    print rows,columns
    for j in range(1,rows+1):
                if parsed.cell(column = 1, row = j).value == 1 :
                        temp = ' '.join(string)
                        sentences[sentence_no] = temp
                        if len(order) == 3:
                            if (order[0] == "S") & (order[1]=="V") & (order[2] == 'O'):
                                SVO.append(sentence_no)
                            elif (order[0] == "S") & (order[1]=="O") & (order[2] == 'V'):
                                SOV.append(sentence_no)
                            elif (order[0] == "V") & (order[1]=="S") & (order[2] == 'O'):
                                VSO.append(sentence_no)
                            elif (order[0] == "V") & (order[1]=="O") & (order[2] == 'S'):
                                VOS.append(sentence_no)
                            elif (order[0] == "O") & (order[1]=="S") & (order[2] == 'V'):
                                OSV.append(sentence_no)
                            elif (order[0] == "O") & (order[1]=="V") & (order[2] == 'S'):
                                OVS.append(sentence_no)
                            else:
                                other.append(sentence_no)
                        else:
                            other.append(sentence_no)
                        sentence_no +=1
                        flags = [False,False,False]
                        order = []
                        string = []

                tags = parsed.cell(column = 3, row = j).value
                word = parsed.cell(column = 2, row = j).value
                if word != None:
                    if (type(word) == long) | (type(word) == float) | (type(word) == int):
                        word = unicode(str(word),"utf-8")
                    string.append(word)#unicode(word,"utf-8"))
                if tags in subjects:
                        if flags[0] == False:
                            order.append("S")
                            flags[0] = True
                elif tags in objects:
                        if flags[1] == False:
                            order.append("O")
                            flags[1] = True
                elif tags in verbs:
                        if flags[2] == False:
                            order.append("V")
                            flags[2] = True
    print sentence_no
    print "SVO:",len(SVO),SVO
    print "SOV",len(SOV),SOV
    print "VSO",len(VSO),VSO
    print "VOS",len(VOS),VOS
    print "OSV",len(OSV),OSV
    print "OVS",len(OVS),OVS
    print "other",len(other)
    print sentences