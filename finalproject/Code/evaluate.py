#!/usr/bin/env python
# -*- coding: utf-8 -*-
import argparse
from itertools import izip
import codecs
from collections import defaultdict
import openpyxl
import sys

sys.stdout = codecs.getwriter('utf-8')(sys.stdout)
sys.stdin = codecs.getreader('utf-8')(sys.stdin)

from PyDictionary import PyDictionary
dictionary = PyDictionary()
PARSER = argparse.ArgumentParser(description="evaluate the reorderings")
PARSER.add_argument("-t", type=str, default="dev", help="file to be evaluated prefix")
PARSER.add_argument("-es", type=str, default="es", help="spanish file")
PARSER.add_argument("-en", type=str, default="en", help="english file")
PARSER.add_argument("-o", type=str, default="output", help="output file")
PARSER.add_argument("-p", type=str, default="POS.xltx", help="output file")
args = PARSER.parse_args()

sys.stdout = codecs.getwriter('utf-8')(sys.stdout)
sys.stdin = codecs.getreader('utf-8')(sys.stdin)

def combine(a, b): return 'data/%s.%s' % (a, b)
def utf8read(file): return codecs.open(file, 'r', 'utf-8')

if __name__ == '__main__':

    if args.t:
        '''
        Finding an alternative for this portion
        for en_sentences,es_sentences in izip(utf8read(combine(args.t,args.en)),utf8read(combine(args.t,args.es))):
            print en_sentences
            new_sent = ""
            new_words = []
            for word in en_sentences.rstrip().lower().split():
                print word
                translation = dictionary.translate(word,"es")
                if translation == None:
                    new_words.append(str(word))
                elif translation in es_sentences.rstrip().lower().split():
                    new_words.append(str(translation))
            new_sent = ' '.join(new_words)
            print new_sent
        '''
        score = defaultdict(int)
        #check for SVO
        wb = openpyxl.load_workbook("data/"+args.p)
        parsed = wb.get_sheet_by_name(args.t)
        sub = defaultdict(defaultdict)
        obj = defaultdict(defaultdict)
        ver = defaultdict(defaultdict)
        sentence_no = 0
        subjects = ["SUBJ"]
        objects = ["DO","IO","OBLC"]
        verbs = ["AUX"]
        rows = parsed.get_highest_row()
        columns = parsed.get_highest_column()
        for i in range(1, int(rows+1)):
            if parsed.cell(column = 1, row = i).value == 1 :
                order = []
                words = []
                flags = [False, False, False]
                subject = ""
                object = ""
                verb = ""
                sentence_no +=1

            if (parsed.cell(column = 8,row = i).value in subjects) & (flags[0]==False):
                    flags[0] = True
                    subject += parsed.cell(column = 2,row = i).value
                    order.append("S")
                    sub[sentence_no] = subject
            elif (parsed.cell(column = 8,row = i).value in objects) & (flags[1]==False):
                    flags[1] = True
                    object += parsed.cell(column = 2,row = i).value
                    order.append("O")
                    obj[sentence_no] = object
            elif (parsed.cell(column = 8,row = i).value in verbs) & (flags[2]==False):
                    flags[2] = True
                    verb += parsed.cell(column = 2,row = i).value
                    order.append("V")
                    ver[sentence_no] = verbs
        total_possible_score = 100 * sentence_no
        sentence_no = 0
        for output_sentences in izip(utf8read(combine(args.t,args.o))):
            sentence_no += 1
            score[sentence_no] = 0
            words = []
            order = []
            for word in output_sentences.rstrip().split():
                if (word == sub[sentence_no]):
                    order.append("S")
                elif(word == obj[sentence_no]):
                    order.append("O")
                elif(word==ver[sentence_no]):
                    order.append("V")
                else:
                    words.append(word)
            if (order[0]=="S")&(order[1]=="V")&(order[2]=="O"):
                score[sentence_no] +=30
        #Check word by word correct ordering
        #Find total