from collections import defaultdict

import openpyxl
wb = openpyxl.load_workbook("data/Book1.xlsx")
parsed = wb.get_sheet_by_name("Sheet1")
rows = parsed.get_highest_row()
columns = parsed.get_highest_column()
string = ""
dictionary  = defaultdict(defaultdict)
for i in range(1,rows):
    print (parsed.cell(column = 2,row = i).value+"::"+parsed.cell(column = 1,row = i).value)
