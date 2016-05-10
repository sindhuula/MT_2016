#!/usr/bin/env python
# -*- coding: utf-8 -*-
#This program will evaluate the accuracy of the reorder code based on 2 criteria:
    #1 the sentence should have the same word order as the english sentence
    #2 the words in the sentence should be aligned as it would be in an english sentence
import argparse
from itertools import izip
from difflib import SequenceMatcher
import codecs
from collections import defaultdict
import openpyxl
import sys
import re
import dictionary
import translate
sys.stdout = codecs.getwriter('utf-8')(sys.stdout)
sys.stdin = codecs.getreader('utf-8')(sys.stdin)
reload(sys)
sys.setdefaultencoding('utf8')
PARSER = argparse.ArgumentParser(description="evaluate the reorderings")
PARSER.add_argument("-t", type=str, default="trial", help="file to be evaluated prefix")
PARSER.add_argument("-es", type=str, default="es", help="spanish file")
PARSER.add_argument("-en", type=str, default="en", help="english file")
PARSER.add_argument("-o", type=str, default="output", help="output file")
PARSER.add_argument("-p", type=str, default="pos.txt", help="output file")
args = PARSER.parse_args()

sys.stdout = codecs.getwriter('utf-8')(sys.stdout)
sys.stdin = codecs.getreader('utf-8')(sys.stdin)

def combine(a, b): return 'data/%s.%s' % (a, b)
def utf8read(file): return codecs.open(file, 'r', 'utf-8')

if __name__ == '__main__':

    if args.t:
        score = defaultdict(int)
        #Check for SVO
        wb = openpyxl.load_workbook("data/postags.xlsx")
        parsed = wb.get_sheet_by_name(args.t)
        sub = defaultdict(defaultdict)
        obj = defaultdict(defaultdict)
        ver = defaultdict(defaultdict)
        sentence_no = 0
        subjects = ["SUBJ"]
        objects = ["DO","IO","OBLC"]
        verbs = ["v"]
        order = []
        words = []
        flags = [False, False, False]
        subject = ""
        object = ""
        verb = ""
        rows = parsed.get_highest_row()
        columns = parsed.get_highest_column()
        for i in range(1, int(rows+1)):
            if parsed.cell(column = 1, row = i).value == 1 :
                if i != 1:
                    order = []
                    words = []
                    flags = [False, False, False]
                    subject = ""
                    object = ""
                    verb = ""
                    sentence_no +=1

            if (parsed.cell(column = 8,row = i).value in subjects) & (flags[0]==False):
                    flags[0] = True
                    subject += parsed.cell(column = 2,row = i).value
                    order.append("S")
                    sub[sentence_no] = subject
            elif (parsed.cell(column = 8,row = i).value in objects) & (flags[1]==False):
                    flags[1] = True
                    object += parsed.cell(column = 2,row = i).value
                    order.append("O")
                    obj[sentence_no] = object
            elif (parsed.cell(column = 4,row = i).value in verbs) & (flags[2]==False):
                    flags[2] = True
                    verb += parsed.cell(column = 2,row = i).value
                    order.append("V")
                    ver[sentence_no] = verb
        total_possible_score = 100 * sentence_no
        sentence_no = 0
        score = defaultdict(float)
        for output_sentences in izip(utf8read(combine(args.t,args.o))):
            sentence_no += 1
            score[sentence_no] = 0
            words = []
            order = []
            for word in output_sentences[0].rstrip().split():
                if (word == sub[sentence_no]):
                    order.append("S")
                elif(word == obj[sentence_no]):
                    order.append("O")
                elif(word==ver[sentence_no]):
                    order.append("V")
                else:
                    words.append(word)
            try:
             if (order[0]=="S")&(order[1]=="V")&(order[2]=="O"):
                score[sentence_no] += 30
             elif (order[0]=="S")&(order[1]=="O")&(order[2]=="V"):
                score[sentence_no] += 25
             elif (order[0]=="O")&(order[1]=="S")&(order[2]=="V"):
                score[sentence_no] += 15
             elif (order[0]=="O")&(order[1]=="V")&(order[2]=="S"):
                score[sentence_no] += 20
             elif (order[0]=="V")&(order[1]=="S")&(order[2]=="O"):
                score[sentence_no] += 10
             elif (order[0]=="V")&(order[1]=="O")&(order[2]=="S"):
                score[sentence_no] += 5
             elif (order[0]=="S")&(order[1]=="O")|(order[0]=="O")&(order[1]=="V"):
                 score[sentence_no] += 15
             else:
                   score[sentence_no] += 0
            except:
                score[sentence_no]+= 0



        #Check for correct word alignment
        dictionary  = dictionary.create_dictionary()
        sent_no = 0


        for es_sentences,en_sentences,o_sentence in zip(open(combine(args.t,args.es)),open(combine(args.t,args.en)),open(combine(args.t,args.o))):
            new_sentence = translate.translate(en_sentences,es_sentences,dictionary)
            sent_no+=1
            difference = SequenceMatcher(None,new_sentence,o_sentence)
            score[sent_no] += difference.ratio() * 70


        total_score = 0.0
        for sents in score:
            total_score += score[sents]
        total_possible = 100.0 * float(sent_no)
        print "Accuracy of computation:",str(total_score/total_possible)
