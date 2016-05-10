#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Outputs a version where the phrases are re-ordered in SVO order
TODO Change the order within the phrase and improve reordering
"""
import openpyxl
import argparse
import codecs
import sys
from collections import defaultdict
from itertools import izip, islice

PARSER = argparse.ArgumentParser(description="Reorder a set of sentences")
PARSER.add_argument("-t", type=str, default="data/train", help="training data prefix")
PARSER.add_argument("-d", type=str, default="data/dev", help="dev data prefix")
PARSER.add_argument("-w", type=str, default="es", help="spanish file")
PARSER.add_argument("-p", type=str, default="data/postags.xlsx", help="tag file")
args = PARSER.parse_args()

sys.stdout = codecs.getwriter('utf-8')(sys.stdout)
sys.stdin = codecs.getreader('utf-8')(sys.stdin)
def sentences():
        with open(combine(args.d,args.w)) as f:
            for pair in f:
                yield [sentence.strip().split() for sentence in pair.split(' ||| ')]

def combine(a, b): return '%s.%s' % (a, b)

if __name__ == '__main__':

    if args.t:
        sentence_tags = defaultdict(list)
        def utf8read(file): return codecs.open(file, 'r', 'utf-8')
        wb = openpyxl.load_workbook(args.p)
        parsed = wb.get_sheet_by_name("train")
        sentence_no = 0
        order = []
        sentences = defaultdict(defaultdict)
        properties = []
        rows = int( parsed.get_highest_row())
        columns = int(parsed.get_highest_column())
        for i in range(1, rows+1):
            if parsed.cell(column = 1, row = i).value == 1 :
                word_props = defaultdict(list)
                sentence_no +=1
                word_props[parsed.cell(column = 2, row = i).value] =  [parsed.cell(column = 4, row = i).value, parsed.cell(column = 5, row = i).value,parsed.cell(column = 8, row = i).value]
                sentences[sentence_no] = word_props
        #Dev Section
        wb = openpyxl.load_workbook(args.p)
        parsed = wb.get_sheet_by_name('dev')
        sentence_no = 1
        order = []
        sentences = defaultdict(defaultdict)
        properties = []
        rows = int( parsed.get_highest_row())
        columns = int(parsed.get_highest_column())
        word_props = defaultdict(list)
        for i in range(1, 62):
            if parsed.cell(column = 1, row = i).value == 1:
                if i == 1:
                    continue
                else:
                    sentences[sentence_no] = word_props
                    sentence_no +=1
                    word_props = defaultdict(list)
            word_props[parsed.cell(column = 2, row = i).value] =  [parsed.cell(column = 4, row = i).value, parsed.cell(column = 5, row = i).value,parsed.cell(column = 8, row = i).value]

        verb_phrase = ""
        subj_phrase = ""
        obj_phrase = ""
        noun_phrase = ""
        subjects = ["SUBJ"]
        objects = ["DO","IO","OBLC"]
        verbs = ["v"]
        print sentences[1]
        for i in range(1, 3):
             final_sent = ""
             temp_phrase = []
             words = sentences[i]
            # for word in words:
            #    print word[2]
            #     if word != None:
            #         if (type(word) == long) | (type(word) == float) | (type(word) == int):
            #             word = unicode(str(word),"utf-8")
            #         temp_phrase.append(word)

            #     if word[2] in subjects:
            #         subj_phrase = ' '.join(temp_phrase)
            #         temp_phrase = []
            #     elif word[2] in objects:
            #         obj_phrase = ' '.join(temp_phrase)
            #         temp_phrase = []
            #     elif word[0] in verbs:
            #         verb_phrase = ' '.join(temp_phrase)
            #         temp_phrase = []
            # final_sent = subj_phrase+' '+verb_phrase+' '+obj_phrase
            # print final_sent
